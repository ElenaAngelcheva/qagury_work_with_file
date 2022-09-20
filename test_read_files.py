from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv


def test_read_pdf():
    reader = PdfReader('resources/files/docs_pytest.pdf')
    number_of_pages = len(reader.pages)
    page = reader.pages[20]
    text = page.extract_text()
    assert number_of_pages == 412
    assert "2.3How to use fixtures" in text


def test_read_xlsx():
    workbook = load_workbook('resources/files/file_example.xlsx')
    sheet = workbook.active
    id = sheet.cell(row=8, column=8).value
    assert id == 3598


def test_read_csv():
    with open('resources/files/username.csv') as file:
        reader = csv.reader(file)
        headers = next(reader)
    assert 'Username' in str(headers)






